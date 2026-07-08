"""Window schedule: manual capture (W2), duplicate (FR-W2), inline config add (FR-W3),
and file import with row-level validation (W3 / FR-W4).

Phase-1 guard-rail (FR-R4): this module only records window geometry; it never
computes any material quantity from it.
"""
import csv
import io
import re

from flask import (
    Blueprint, abort, flash, redirect, render_template, request, send_file, url_for, Response,
)

from .db import get_db
from .services import get_project_or_404, window_summary

bp = Blueprint('windows', __name__)

TEMPLATE_HEADERS = ['Opening Ref', 'Location', 'Config Code', 'Width mm', 'Height mm', 'Qty', 'Remarks']


def _active_configs(db):
    return db.execute(
        'SELECT * FROM window_config_types WHERE active = 1 ORDER BY code COLLATE NOCASE'
    ).fetchall()


# --- Manual capture ---------------------------------------------------------

@bp.route('/projects/<int:project_id>/windows')
def list_windows(project_id):
    db = get_db()
    project = get_project_or_404(project_id)
    lines = db.execute(
        """
        SELECT w.*, c.code AS config_code, c.name AS config_name
        FROM window_lines w JOIN window_config_types c ON c.id = w.config_type_id
        WHERE w.project_id = ? ORDER BY w.opening_ref COLLATE NOCASE, w.id
        """,
        (project_id,),
    ).fetchall()
    imports = db.execute(
        'SELECT id, filename, status, rows_imported, rows_rejected, uploaded_at'
        ' FROM import_files WHERE project_id = ? ORDER BY uploaded_at DESC, id DESC',
        (project_id,),
    ).fetchall()
    return render_template(
        'windows/list.html',
        project=project, lines=lines, configs=_active_configs(db),
        summary=window_summary(db, project_id), imports=imports,
    )


def _window_form(db, project_id):
    """Parse the window-line form. Handles the inline 'add new config' fields (FR-W3).

    Returns (values, errors).
    """
    errors = []
    opening_ref = request.form.get('opening_ref', '').strip()
    location = request.form.get('location', '').strip()
    tostem_order_ref = request.form.get('tostem_order_ref', '').strip()
    remarks = request.form.get('remarks', '').strip()

    config_type_id = None
    new_code = request.form.get('new_config_code', '').strip()
    if new_code:
        existing = db.execute(
            'SELECT id FROM window_config_types WHERE code = ?', (new_code,)
        ).fetchone()
        if existing:
            config_type_id = existing['id']
        else:
            new_name = request.form.get('new_config_name', '').strip() or new_code
            tracks = request.form.get('new_config_tracks', '').strip()
            panels = request.form.get('new_config_panels', '').strip()
            cur = db.execute(
                'INSERT INTO window_config_types (code, name, tracks, panels) VALUES (?, ?, ?, ?)',
                (new_code, new_name,
                 int(tracks) if tracks.isdigit() else None,
                 int(panels) if panels.isdigit() else None),
            )
            config_type_id = cur.lastrowid
            flash(f"Configuration type '{new_code}' created.", 'success')
    else:
        raw = request.form.get('config_type_id', '').strip()
        if raw.isdigit():
            row = db.execute(
                'SELECT id FROM window_config_types WHERE id = ?', (int(raw),)
            ).fetchone()
            if row:
                config_type_id = row['id']

    def positive_number(field, label):
        raw = request.form.get(field, '').strip()
        try:
            val = float(raw)
        except ValueError:
            errors.append(f'{label} must be a number.')
            return None
        if val <= 0:
            errors.append(f'{label} must be greater than zero.')
            return None
        return val

    width = positive_number('width_mm', 'Width (mm)')
    height = positive_number('height_mm', 'Height (mm)')
    qty_raw = request.form.get('quantity', '').strip()
    quantity = int(qty_raw) if qty_raw.isdigit() and int(qty_raw) > 0 else None
    if quantity is None:
        errors.append('Quantity must be a positive whole number.')
    if not opening_ref:
        errors.append('Opening ref is required.')
    if config_type_id is None:
        errors.append('Configuration type is required (pick one or add a new one inline).')

    return {
        'opening_ref': opening_ref, 'location': location, 'config_type_id': config_type_id,
        'width_mm': width, 'height_mm': height, 'quantity': quantity,
        'tostem_order_ref': tostem_order_ref, 'remarks': remarks,
    }, errors


@bp.route('/projects/<int:project_id>/windows/add', methods=('POST',))
def add_window(project_id):
    db = get_db()
    get_project_or_404(project_id)
    f, errors = _window_form(db, project_id)
    if errors:
        db.commit()  # keep any inline-created config type even if the line is invalid
        for e in errors:
            flash(e, 'error')
    else:
        db.execute(
            'INSERT INTO window_lines (project_id, opening_ref, location, config_type_id,'
            " width_mm, height_mm, quantity, tostem_order_ref, source, remarks)"
            " VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Manual', ?)",
            (project_id, f['opening_ref'], f['location'], f['config_type_id'],
             f['width_mm'], f['height_mm'], f['quantity'], f['tostem_order_ref'], f['remarks']),
        )
        db.commit()
        flash('Window line added.', 'success')
    return redirect(url_for('windows.list_windows', project_id=project_id))


def _get_line_or_404(db, project_id, line_id):
    line = db.execute(
        'SELECT * FROM window_lines WHERE id = ? AND project_id = ?', (line_id, project_id)
    ).fetchone()
    if line is None:
        abort(404)
    return line


@bp.route('/projects/<int:project_id>/windows/<int:line_id>/edit', methods=('GET', 'POST'))
def edit_window(project_id, line_id):
    db = get_db()
    project = get_project_or_404(project_id)
    line = _get_line_or_404(db, project_id, line_id)
    if request.method == 'POST':
        f, errors = _window_form(db, project_id)
        if errors:
            db.commit()  # keep any inline-created config type even if the line is invalid
            for e in errors:
                flash(e, 'error')
        else:
            db.execute(
                'UPDATE window_lines SET opening_ref = ?, location = ?, config_type_id = ?,'
                ' width_mm = ?, height_mm = ?, quantity = ?, tostem_order_ref = ?, remarks = ?,'
                ' updated_at = CURRENT_TIMESTAMP WHERE id = ?',
                (f['opening_ref'], f['location'], f['config_type_id'], f['width_mm'],
                 f['height_mm'], f['quantity'], f['tostem_order_ref'], f['remarks'], line_id),
            )
            db.commit()
            flash('Window line updated.', 'success')
            return redirect(url_for('windows.list_windows', project_id=project_id))
    return render_template(
        'windows/form.html', project=project, line=line, configs=_active_configs(db),
    )


@bp.route('/projects/<int:project_id>/windows/<int:line_id>/delete', methods=('POST',))
def delete_window(project_id, line_id):
    db = get_db()
    _get_line_or_404(db, project_id, line_id)
    db.execute('DELETE FROM window_lines WHERE id = ?', (line_id,))
    db.commit()
    flash('Window line deleted.', 'success')
    return redirect(url_for('windows.list_windows', project_id=project_id))


@bp.route('/projects/<int:project_id>/windows/<int:line_id>/duplicate', methods=('POST',))
def duplicate_window(project_id, line_id):
    db = get_db()
    line = _get_line_or_404(db, project_id, line_id)
    cur = db.execute(
        'INSERT INTO window_lines (project_id, opening_ref, location, config_type_id, width_mm,'
        " height_mm, quantity, tostem_order_ref, source, remarks)"
        " VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'Manual', ?)",
        (project_id, line['opening_ref'] + ' (copy)', line['location'], line['config_type_id'],
         line['width_mm'], line['height_mm'], line['quantity'],
         line['tostem_order_ref'], line['remarks']),
    )
    db.commit()
    flash('Window line duplicated — adjust the opening ref below.', 'success')
    return redirect(url_for('windows.edit_window', project_id=project_id, line_id=cur.lastrowid))


# --- Import (W3 / FR-W4) ------------------------------------------------------

def _normalize_header(h):
    return re.sub(r'[^a-z]', '', str(h or '').lower())


HEADER_KEYS = {
    'openingref': 'opening_ref',
    'location': 'location',
    'configcode': 'config_code',
    'widthmm': 'width_mm',
    'heightmm': 'height_mm',
    'qty': 'quantity',
    'quantity': 'quantity',
    'remarks': 'remarks',
}
REQUIRED_COLUMNS = ('opening_ref', 'config_code', 'width_mm', 'height_mm', 'quantity')


def parse_import_content(filename, content):
    """Return (rows, error). rows = list of dicts keyed by canonical column names."""
    name = filename.lower()
    if name.endswith('.csv'):
        try:
            text = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = content.decode('latin-1')
        raw_rows = list(csv.reader(io.StringIO(text)))
    elif name.endswith('.xlsx'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.worksheets[0]
            raw_rows = [['' if c is None else c for c in row] for row in ws.iter_rows(values_only=True)]
        except Exception:
            return None, 'Could not read the XLSX file. Please use the app template.'
    else:
        return None, 'Unsupported file type. Upload a .csv or .xlsx file using the app template.'

    raw_rows = [r for r in raw_rows if any(str(c).strip() for c in r)]
    if not raw_rows:
        return None, 'The file is empty.'

    header = raw_rows[0]
    col_map = {}
    for idx, cell in enumerate(header):
        key = HEADER_KEYS.get(_normalize_header(cell))
        if key and key not in col_map:
            col_map[key] = idx
    missing = [c for c in REQUIRED_COLUMNS if c not in col_map]
    if missing:
        return None, (
            'The file does not match the template. Missing columns: '
            + ', '.join(missing).replace('_', ' ')
            + '. Download the template and try again.'
        )

    rows = []
    for i, raw in enumerate(raw_rows[1:], start=2):
        def cell(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(raw):
                return ''
            return str(raw[idx]).strip()
        rows.append({
            'row_no': i,
            'opening_ref': cell('opening_ref'),
            'location': cell('location'),
            'config_code': cell('config_code'),
            'width_mm': cell('width_mm'),
            'height_mm': cell('height_mm'),
            'quantity': cell('quantity'),
            'remarks': cell('remarks'),
        })
    return rows, None


def validate_import_rows(db, rows):
    """Row-level validation per FR-W4. Returns (validated_rows, unknown_codes)."""
    configs = {
        r['code'].lower(): r
        for r in db.execute('SELECT * FROM window_config_types WHERE active = 1').fetchall()
    }
    unknown_codes = []
    validated = []
    for row in rows:
        errors = []
        r = dict(row)
        if not r['opening_ref']:
            errors.append('Opening Ref is required')

        code = r['config_code']
        r['config_type_id'] = None
        if not code:
            errors.append('Config Code is required')
        elif code.lower() in configs:
            r['config_type_id'] = configs[code.lower()]['id']
        else:
            errors.append(f"Unknown config code '{code}'")
            if code not in unknown_codes:
                unknown_codes.append(code)

        for field, label in (('width_mm', 'Width mm'), ('height_mm', 'Height mm')):
            raw = r[field]
            try:
                val = float(raw)
                if val <= 0:
                    errors.append(f'{label} must be greater than zero')
                    val = None
            except ValueError:
                errors.append(f'{label} is missing or not a number')
                val = None
            r[field + '_value'] = val

        qraw = r['quantity']
        try:
            qf = float(qraw)
            if qf <= 0 or qf != int(qf):
                errors.append('Qty must be a positive whole number')
                r['quantity_value'] = None
            else:
                r['quantity_value'] = int(qf)
        except ValueError:
            errors.append('Qty is missing or not a number')
            r['quantity_value'] = None

        r['errors'] = errors
        validated.append(r)
    return validated, unknown_codes


@bp.route('/windows/import-template.csv')
def import_template():
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerow(['W-101', '3rd floor, Bedroom 2', '3T3P', '2400', '1500', '2', ''])
    return Response(
        out.getvalue(), mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=window-schedule-template.csv'},
    )


@bp.route('/windows/import-template.xlsx')
def import_template_xlsx():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Window Schedule'
    ws.append(TEMPLATE_HEADERS)
    ws.append(['W-101', '3rd floor, Bedroom 2', '3T3P', 2400, 1500, 2, ''])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name='window-schedule-template.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@bp.route('/projects/<int:project_id>/windows/import', methods=('GET', 'POST'))
def import_upload(project_id):
    db = get_db()
    project = get_project_or_404(project_id)
    if request.method == 'POST':
        file = request.files.get('file')
        if file is None or not file.filename:
            flash('Choose a CSV or XLSX file to upload.', 'error')
        else:
            content = file.read()
            rows, error = parse_import_content(file.filename, content)
            if error:
                flash(error, 'error')
            else:
                cur = db.execute(
                    'INSERT INTO import_files (project_id, filename, content) VALUES (?, ?, ?)',
                    (project_id, file.filename, content),
                )
                db.commit()
                return redirect(url_for(
                    'windows.import_review', project_id=project_id, import_id=cur.lastrowid,
                ))
    return render_template('windows/import_upload.html', project=project)


def _get_import_or_404(db, project_id, import_id):
    imp = db.execute(
        'SELECT * FROM import_files WHERE id = ? AND project_id = ?', (import_id, project_id)
    ).fetchone()
    if imp is None:
        abort(404)
    return imp


@bp.route('/projects/<int:project_id>/windows/import/<int:import_id>/review')
def import_review(project_id, import_id):
    db = get_db()
    project = get_project_or_404(project_id)
    imp = _get_import_or_404(db, project_id, import_id)
    if imp['status'] != 'Pending':
        flash('This import has already been processed.', 'error')
        return redirect(url_for('windows.list_windows', project_id=project_id))
    rows, error = parse_import_content(imp['filename'], imp['content'])
    if error:
        flash(error, 'error')
        return redirect(url_for('windows.import_upload', project_id=project_id))
    validated, unknown_codes = validate_import_rows(db, rows)
    valid_count = sum(1 for r in validated if not r['errors'])
    return render_template(
        'windows/import_review.html',
        project=project, imp=imp, rows=validated, unknown_codes=unknown_codes,
        valid_count=valid_count, invalid_count=len(validated) - valid_count,
    )


@bp.route('/projects/<int:project_id>/windows/import/<int:import_id>/confirm', methods=('POST',))
def import_confirm(project_id, import_id):
    db = get_db()
    get_project_or_404(project_id)
    imp = _get_import_or_404(db, project_id, import_id)
    if imp['status'] != 'Pending':
        flash('This import has already been processed.', 'error')
        return redirect(url_for('windows.list_windows', project_id=project_id))
    rows, error = parse_import_content(imp['filename'], imp['content'])
    if error:
        flash(error, 'error')
        return redirect(url_for('windows.import_upload', project_id=project_id))
    validated, _ = validate_import_rows(db, rows)
    imported = 0
    rejected = 0
    for r in validated:
        if r['errors']:
            rejected += 1
            continue
        db.execute(
            'INSERT INTO window_lines (project_id, opening_ref, location, config_type_id,'
            " width_mm, height_mm, quantity, source, remarks)"
            " VALUES (?, ?, ?, ?, ?, ?, ?, 'Imported', ?)",
            (project_id, r['opening_ref'], r['location'], r['config_type_id'],
             r['width_mm_value'], r['height_mm_value'], r['quantity_value'], r['remarks']),
        )
        imported += 1
    db.execute(
        "UPDATE import_files SET status = 'Imported', rows_imported = ?, rows_rejected = ?"
        ' WHERE id = ?',
        (imported, rejected, import_id),
    )
    db.commit()
    msg = f'Imported {imported} window line(s).'
    if rejected:
        msg += f' {rejected} invalid row(s) were rejected.'
    flash(msg, 'success')
    return redirect(url_for('windows.list_windows', project_id=project_id))


@bp.route('/projects/<int:project_id>/windows/import/<int:import_id>/discard', methods=('POST',))
def import_discard(project_id, import_id):
    db = get_db()
    get_project_or_404(project_id)
    imp = _get_import_or_404(db, project_id, import_id)
    if imp['status'] == 'Pending':
        db.execute("UPDATE import_files SET status = 'Discarded' WHERE id = ?", (import_id,))
        db.commit()
        flash('Import discarded. The uploaded file is kept for audit.', 'success')
    return redirect(url_for('windows.list_windows', project_id=project_id))


@bp.route('/projects/<int:project_id>/windows/import/<int:import_id>/download')
def import_download(project_id, import_id):
    db = get_db()
    get_project_or_404(project_id)
    imp = _get_import_or_404(db, project_id, import_id)
    return send_file(
        io.BytesIO(imp['content']), as_attachment=True, download_name=imp['filename'],
    )
