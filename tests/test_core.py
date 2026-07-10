"""Core-logic tests keyed to the PRD's functional requirements and acceptance criteria."""
import io
import re
import time
from datetime import date

from conftest import (
    ADMIN_EMAIL, ADMIN_PASSWORD, add_requirement, add_tool, create_po, execute, material_id,
    query, query_one,
)

DRAWING_PDF = "/root/.claude/uploads/6ef23547-c079-5885-9d05-22062850ef2d/d5d2d9d6-TG_GROUP_GLASS_SIZES_1.pdf"


# --- Auth (FR-G1) -----------------------------------------------------------

def test_login_required(client):
    resp = client.get('/projects/')
    assert resp.status_code == 302
    assert '/login' in resp.headers['Location']


def test_login_and_logout(client):
    resp = client.post('/login', data={'email': ADMIN_EMAIL, 'password': 'wrong'})
    assert b'Invalid email or password' in resp.data
    resp = client.post('/login', data={'email': ADMIN_EMAIL, 'password': ADMIN_PASSWORD})
    assert resp.status_code == 302
    assert client.get('/projects/').status_code == 200
    client.post('/logout')
    assert client.get('/projects/').status_code == 302


def test_session_timeout(app, auth):
    with auth.session_transaction() as sess:
        sess['last_active'] = time.time() - (app.config['SESSION_TIMEOUT_MINUTES'] * 60 + 10)
    resp = auth.get('/projects/')
    assert resp.status_code == 302
    assert '/login' in resp.headers['Location']


# --- Catalog seed (FR-C2) and deactivation (FR-C3) ---------------------------

def test_seed_materials(app):
    rows = query(app, 'SELECT name, uom, supply_source FROM materials ORDER BY id')
    assert len(rows) == 9
    names = {r['name'] for r in rows}
    assert names == {'Glass', 'Silicone', 'Packers', 'Screws', 'Masking Tape',
                     'Gasket', 'Safety Shoes', 'Helmets', 'Jackets'}
    gasket = query_one(app, "SELECT * FROM materials WHERE name = 'Gasket'")
    assert gasket['supply_source'] == 'Tostem-supplied'
    assert gasket['uom'] == 'm'


def test_material_deactivation_hidden_from_forms(app, auth, project):
    mid = material_id(app, 'Packers')
    line = add_requirement(auth, project['id'], mid, 50, app=app)
    auth.post(f'/materials/{mid}/toggle-active')
    # Hidden from the new-entry form...
    resp = auth.get(f"/projects/{project['id']}/requirements/")
    assert 'Packers (nos' not in resp.get_data(as_text=True)
    # ...but preserved in historical records (FR-C3).
    resp_text = resp.get_data(as_text=True)
    assert 'Packers' in resp_text
    assert query_one(app, 'SELECT * FROM requirement_lines WHERE id = ?', (line['id'],))


def test_new_material_addable(app, auth):
    # AC2: a new material (e.g. "Backer Rod") can be added with no code change.
    auth.post('/materials/new', data={
        'name': 'Backer Rod', 'category': 'Window Consumable', 'uom': 'm',
        'supply_source': 'Vendor-procured', 'default_vendor_id': '', 'notes': '',
    })
    row = query_one(app, "SELECT * FROM materials WHERE name = 'Backer Rod'")
    assert row is not None and row['active'] == 1


# --- Window schedule (FR-W1..W5) ---------------------------------------------

def _config_id(app, code='3T3P'):
    return query_one(app, 'SELECT id FROM window_config_types WHERE code = ?', (code,))['id']


def test_window_line_validation(app, auth, project):
    resp = auth.post(f"/projects/{project['id']}/windows/add", data={
        'opening_ref': 'W-101', 'config_type_id': str(_config_id(app)),
        'width_mm': 'abc', 'height_mm': '1500', 'quantity': '2',
    }, follow_redirects=True)
    assert b'Width (mm) must be a number' in resp.data
    assert query(app, 'SELECT * FROM window_lines') == []


def test_window_add_duplicate_and_summary(app, auth, project):
    pid = project['id']
    auth.post(f'/projects/{pid}/windows/add', data={
        'opening_ref': 'W-101', 'location': 'Bedroom 2', 'config_type_id': str(_config_id(app)),
        'width_mm': '2400', 'height_mm': '1500', 'quantity': '2',
    })
    line = query_one(app, 'SELECT * FROM window_lines ORDER BY id DESC LIMIT 1')
    assert line['source'] == 'Manual'
    auth.post(f"/projects/{pid}/windows/{line['id']}/duplicate")
    lines = query(app, 'SELECT * FROM window_lines ORDER BY id')
    assert len(lines) == 2
    assert lines[1]['opening_ref'] == 'W-101 (copy)'
    resp = auth.get(f'/projects/{pid}/windows')
    text = resp.get_data(as_text=True)
    # Summary strip (FR-W5): 2 opening lines, 4 windows (lines × qty).
    assert re.search(r'stat-value">2</span>\s*<span class="stat-label">openings', text)
    assert re.search(r'stat-value">4</span>\s*<span class="stat-label">windows', text)


def test_inline_config_creation(app, auth, project):
    # FR-W3 / AC2: add a new configuration type (2T2P) from the window form.
    auth.post(f"/projects/{project['id']}/windows/add", data={
        'opening_ref': 'W-201', 'config_type_id': '',
        'new_config_code': '2T2P', 'new_config_name': '2 Track 2 Panel Sliding',
        'new_config_tracks': '2', 'new_config_panels': '2',
        'width_mm': '1200', 'height_mm': '1200', 'quantity': '1',
    })
    cfg = query_one(app, "SELECT * FROM window_config_types WHERE code = '2T2P'")
    assert cfg is not None
    line = query_one(app, 'SELECT * FROM window_lines ORDER BY id DESC LIMIT 1')
    assert line['config_type_id'] == cfg['id']


def test_import_flow(app, auth, project):
    # W3 / FR-W4 / AC1: import with one invalid row and one unknown config code.
    pid = project['id']
    csv_content = (
        'Opening Ref,Location,Config Code,Width mm,Height mm,Qty,Remarks\n'
        'W-101,Floor 1,3T3P,2400,1500,2,\n'
        'W-102,Floor 2,2T2P,1200,1200,1,corner window\n'
        'W-103,Floor 3,3T3P,abc,1500,1,\n'
    ).encode()
    resp = auth.post(
        f'/projects/{pid}/windows/import',
        data={'file': (io.BytesIO(csv_content), 'schedule.csv')},
        content_type='multipart/form-data',
    )
    assert resp.status_code == 302
    review_url = resp.headers['Location']
    text = auth.get(review_url).get_data(as_text=True)
    assert "Unknown config code &#39;2T2P&#39;" in text or "Unknown config code '2T2P'" in text
    assert 'Create config' in text
    assert 'not a number' in text  # W-103 width

    # One-click create of the missing config, then the row revalidates as OK.
    auth.post('/configs/quick-create', data={'code': '2T2P', 'next': review_url})
    text = auth.get(review_url).get_data(as_text=True)
    assert '2</strong> valid row(s)' in text

    imp = query_one(app, 'SELECT * FROM import_files ORDER BY id DESC LIMIT 1')
    auth.post(f"/projects/{pid}/windows/import/{imp['id']}/confirm")
    imp = query_one(app, 'SELECT * FROM import_files WHERE id = ?', (imp['id'],))
    assert imp['status'] == 'Imported'
    assert imp['rows_imported'] == 2
    assert imp['rows_rejected'] == 1
    lines = query(app, 'SELECT * FROM window_lines WHERE project_id = ? ORDER BY id', (pid,))
    assert [l['opening_ref'] for l in lines] == ['W-101', 'W-102']
    assert all(l['source'] == 'Imported' for l in lines)

    # Original file retained for audit (§7.9) and downloadable.
    resp = auth.get(f"/projects/{pid}/windows/import/{imp['id']}/download")
    assert resp.data == csv_content
    # Double-confirm is blocked.
    auth.post(f"/projects/{pid}/windows/import/{imp['id']}/confirm")
    assert len(query(app, 'SELECT * FROM window_lines WHERE project_id = ?', (pid,))) == 2


def test_import_rejects_wrong_template(auth, project):
    resp = auth.post(
        f"/projects/{project['id']}/windows/import",
        data={'file': (io.BytesIO(b'Foo,Bar\n1,2\n'), 'bad.csv')},
        content_type='multipart/form-data', follow_redirects=True,
    )
    assert b'does not match the template' in resp.data


# --- Requirements (FR-R1..R4) ---------------------------------------------------

def test_requirement_add_and_multiple_lines_same_material(app, auth, project):
    mid = material_id(app, 'Silicone')
    add_requirement(auth, project['id'], mid, 40, note='lot 1', app=app)
    add_requirement(auth, project['id'], mid, 60, note='lot 2', app=app)
    rows = query(app, 'SELECT * FROM requirement_lines WHERE project_id = ?', (project['id'],))
    assert len(rows) == 2  # FR-R3
    assert all(r['uom'] == 'tubes' for r in rows)  # uom inherited from material
    resp = auth.get(f"/projects/{project['id']}")
    text = resp.get_data(as_text=True)
    assert '<td class="num cell-strong">100</td>' in text  # coverage aggregates to 100 required


def test_requirement_rejects_bad_qty(app, auth, project):
    mid = material_id(app, 'Glass')
    resp = auth.post(f"/projects/{project['id']}/requirements/add", data={
        'material_id': str(mid), 'required_qty': '-5', 'basis_note': '',
    }, follow_redirects=True)
    assert b'positive number' in resp.data
    assert query(app, 'SELECT * FROM requirement_lines') == []


def test_approved_lines_locked(app, auth, project):
    # FR-R2 / W5: approved lines lock; editing requires un-approving first.
    mid = material_id(app, 'Glass')
    line = add_requirement(auth, project['id'], mid, 120, approve=True, app=app)
    assert line['status'] == 'Approved'

    resp = auth.get(
        f"/projects/{project['id']}/requirements/{line['id']}/edit", follow_redirects=True)
    assert b'locked' in resp.data
    auth.post(f"/projects/{project['id']}/requirements/{line['id']}/edit", data={
        'material_id': str(mid), 'required_qty': '999', 'basis_note': '',
    })
    assert query_one(app, 'SELECT required_qty FROM requirement_lines WHERE id = ?',
                     (line['id'],))['required_qty'] == 120

    resp = auth.post(
        f"/projects/{project['id']}/requirements/{line['id']}/delete", follow_redirects=True)
    assert b'locked' in resp.data
    assert query_one(app, 'SELECT * FROM requirement_lines WHERE id = ?', (line['id'],))

    auth.post(f"/projects/{project['id']}/requirements/{line['id']}/unapprove")
    auth.post(f"/projects/{project['id']}/requirements/{line['id']}/edit", data={
        'material_id': str(mid), 'required_qty': '150', 'basis_note': '',
    })
    assert query_one(app, 'SELECT required_qty FROM requirement_lines WHERE id = ?',
                     (line['id'],))['required_qty'] == 150


# --- Purchase orders (FR-O1..O6) --------------------------------------------------

def test_po_creation_groups_and_numbering(app, auth, project, vendor):
    mid = material_id(app, 'Silicone')
    line = add_requirement(auth, project['id'], mid, 100, approve=True, app=app)
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    assert po['po_number'] == f'SS-PO-{date.today().year}-001'  # FR-O3
    assert po['status'] == 'Draft'
    pol = query_one(app, 'SELECT * FROM po_lines WHERE po_id = ?', (po['id'],))
    assert pol['qty'] == 100
    assert pol['requirement_line_id'] == line['id']  # FR-O1 traceability

    line2 = add_requirement(auth, project['id'], material_id(app, 'Glass'), 50,
                            approve=True, app=app)
    po2 = create_po(auth, app, project['id'], vendor['id'], [line2['id']])
    assert po2['po_number'] == f'SS-PO-{date.today().year}-002'


def test_draft_po_from_unapproved_line_rejected(app, auth, project, vendor):
    mid = material_id(app, 'Silicone')
    line = add_requirement(auth, project['id'], mid, 100, app=app)  # Draft
    resp = auth.post(f"/projects/{project['id']}/purchase-orders/from-requirements",
                     data={'line_ids': [str(line['id'])]}, follow_redirects=True)
    assert b'Only Approved requirement lines' in resp.data
    assert query(app, 'SELECT * FROM purchase_orders') == []


def test_fr_o2_tostem_rejected_from_po_selection(app, auth, project, vendor):
    # AC4: attempting to put Gasket on a PO is blocked with a message pointing
    # to the Tostem Requisition.
    gasket = material_id(app, 'Gasket')
    line = add_requirement(auth, project['id'], gasket, 480, approve=True, app=app)
    resp = auth.post(f"/projects/{project['id']}/purchase-orders/from-requirements",
                     data={'line_ids': [str(line['id'])]}, follow_redirects=True)
    assert b'Tostem Requisition' in resp.data
    assert b'cannot be placed on a vendor purchase order' in resp.data
    assert query(app, 'SELECT * FROM purchase_orders') == []


def test_fr_o2_tostem_rejected_from_draft_po_add_line(app, auth, project, vendor):
    mid = material_id(app, 'Silicone')
    line = add_requirement(auth, project['id'], mid, 10, approve=True, app=app)
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    gasket = material_id(app, 'Gasket')
    resp = auth.post(f"/purchase-orders/{po['id']}/lines/add", data={
        'material_id': str(gasket), 'qty': '480', 'rate': '',
    }, follow_redirects=True)
    assert b'Tostem Requisition' in resp.data
    assert len(query(app, 'SELECT * FROM po_lines WHERE po_id = ?', (po['id'],))) == 1


def test_po_lifecycle_partial_receipt_and_coverage(app, auth, project, vendor):
    # W8 / FR-S1 / AC6: partial receipt updates PO status and on-hand; coverage
    # reconciles exactly with the ledger.
    pid = project['id']
    mid = material_id(app, 'Packers')
    line = add_requirement(auth, pid, mid, 10, approve=True, app=app)
    po = create_po(auth, app, pid, vendor['id'], [line['id']])
    pol = query_one(app, 'SELECT * FROM po_lines WHERE po_id = ?', (po['id'],))

    # Draft POs do not count as Ordered; receive before issue is refused.
    resp = auth.get(f"/purchase-orders/{po['id']}/receive", follow_redirects=True)
    assert b'issued PO' in resp.data

    auth.post(f"/purchase-orders/{po['id']}/issue")
    assert query_one(app, 'SELECT status FROM purchase_orders WHERE id = ?',
                     (po['id'],))['status'] == 'Issued'

    auth.post(f"/purchase-orders/{po['id']}/receive", data={
        'date': '2026-07-08', 'note': 'first lot', f"qty_{pol['id']}": '4',
    })
    assert query_one(app, 'SELECT status FROM purchase_orders WHERE id = ?',
                     (po['id'],))['status'] == 'Partially Received'

    cov = _coverage_row(app, auth, pid, 'Packers')
    assert (cov['required'], cov['ordered'], cov['received'], cov['on_hand']) == \
        ('10', '10', '4', '4')

    auth.post(f"/purchase-orders/{po['id']}/receive", data={
        'date': '2026-07-09', f"qty_{pol['id']}": '6',
    })
    assert query_one(app, 'SELECT status FROM purchase_orders WHERE id = ?',
                     (po['id'],))['status'] == 'Received'
    txns = query(app, "SELECT * FROM stock_transactions WHERE type = 'Receipt'")
    assert sum(t['qty'] for t in txns) == 10
    assert all(t['po_line_id'] == pol['id'] for t in txns)

    # Close after full receipt (FR-O4).
    auth.post(f"/purchase-orders/{po['id']}/close")
    assert query_one(app, 'SELECT status FROM purchase_orders WHERE id = ?',
                     (po['id'],))['status'] == 'Closed'


def _coverage_row(app, auth, project_id, material_name):
    """Read the coverage row for a material straight from the services layer."""
    from mpapp.services import coverage_rows
    from mpapp.db import get_db
    with app.app_context():
        rows = coverage_rows(get_db(), project_id)
    for r in rows:
        if r['name'] == material_name:
            return {k: (('%g' % v) if isinstance(v, (int, float)) else v)
                    for k, v in r.items()}
    raise AssertionError(f'{material_name} not in coverage')


def test_over_receipt_requires_confirmation(app, auth, project, vendor):
    # FR-S1: over-receipt permitted only after an explicit warning confirmation.
    pid = project['id']
    line = add_requirement(auth, pid, material_id(app, 'Screws'), 5, approve=True, app=app)
    po = create_po(auth, app, pid, vendor['id'], [line['id']])
    pol = query_one(app, 'SELECT * FROM po_lines WHERE po_id = ?', (po['id'],))
    auth.post(f"/purchase-orders/{po['id']}/issue")

    resp = auth.post(f"/purchase-orders/{po['id']}/receive", data={
        'date': '2026-07-08', f"qty_{pol['id']}": '7',
    })
    assert b'Over-receipt warning' in resp.data
    assert query(app, 'SELECT * FROM stock_transactions') == []  # nothing posted yet

    auth.post(f"/purchase-orders/{po['id']}/receive", data={
        'date': '2026-07-08', f"qty_{pol['id']}": '7', 'confirm_over': '1',
    })
    txns = query(app, 'SELECT * FROM stock_transactions')
    assert len(txns) == 1 and txns[0]['qty'] == 7
    assert query_one(app, 'SELECT status FROM purchase_orders WHERE id = ?',
                     (po['id'],))['status'] == 'Received'


def test_cancel_only_before_receipt(app, auth, project, vendor):
    # FR-O4: Cancelled allowed only before any receipt.
    pid = project['id']
    line = add_requirement(auth, pid, material_id(app, 'Glass'), 20, approve=True, app=app)
    po = create_po(auth, app, pid, vendor['id'], [line['id']])
    pol = query_one(app, 'SELECT * FROM po_lines WHERE po_id = ?', (po['id'],))
    auth.post(f"/purchase-orders/{po['id']}/issue")
    auth.post(f"/purchase-orders/{po['id']}/receive", data={
        'date': '2026-07-08', f"qty_{pol['id']}": '5',
    })
    resp = auth.post(f"/purchase-orders/{po['id']}/cancel", follow_redirects=True)
    assert b'can no longer be cancelled' in resp.data
    assert query_one(app, 'SELECT status FROM purchase_orders WHERE id = ?',
                     (po['id'],))['status'] == 'Partially Received'


def test_cancel_draft_po(app, auth, project, vendor):
    line = add_requirement(auth, project['id'], material_id(app, 'Glass'), 20,
                           approve=True, app=app)
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    auth.post(f"/purchase-orders/{po['id']}/cancel")
    assert query_one(app, 'SELECT status FROM purchase_orders WHERE id = ?',
                     (po['id'],))['status'] == 'Cancelled'


def test_issued_po_locked_for_editing(app, auth, project, vendor):
    line = add_requirement(auth, project['id'], material_id(app, 'Glass'), 20,
                           approve=True, app=app)
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    pol = query_one(app, 'SELECT * FROM po_lines WHERE po_id = ?', (po['id'],))
    auth.post(f"/purchase-orders/{po['id']}/issue")
    resp = auth.post(f"/purchase-orders/{po['id']}/lines/{pol['id']}/update",
                     data={'qty': '99', 'rate': ''}, follow_redirects=True)
    assert b'Only Draft POs can be edited' in resp.data
    assert query_one(app, 'SELECT qty FROM po_lines WHERE id = ?', (pol['id'],))['qty'] == 20


def test_po_totals_only_with_rates(app, auth, project, vendor):
    # FR-O6: totals appear only when rates are present.
    line = add_requirement(auth, project['id'], material_id(app, 'Glass'), 20,
                           approve=True, app=app)
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    pol = query_one(app, 'SELECT * FROM po_lines WHERE po_id = ?', (po['id'],))
    text = auth.get(f"/purchase-orders/{po['id']}/print").get_data(as_text=True)
    assert 'Total' not in text
    auth.post(f"/purchase-orders/{po['id']}/lines/{pol['id']}/update",
              data={'qty': '20', 'rate': '55.5'})
    text = auth.get(f"/purchase-orders/{po['id']}/print").get_data(as_text=True)
    assert 'Total' in text and '1,110.00' in text


# --- Stock (FR-S2..S5) -------------------------------------------------------------

def test_direct_receipt_without_po(app, auth, project):
    # FR-S2: Tostem gaskets arriving with the windows, no PO.
    gasket = material_id(app, 'Gasket')
    auth.post(f"/projects/{project['id']}/stock/receipt", data={
        'material_id': str(gasket), 'qty': '480', 'date': '2026-07-08',
        'note': 'arrived with window shipment',
    })
    txn = query_one(app, 'SELECT * FROM stock_transactions ORDER BY id DESC LIMIT 1')
    assert txn['type'] == 'Receipt' and txn['qty'] == 480 and txn['po_line_id'] is None


def test_issue_blocked_beyond_on_hand(app, auth, project):
    # FR-S5 / AC8.
    pid = project['id']
    gasket = material_id(app, 'Gasket')
    auth.post(f'/projects/{pid}/stock/receipt', data={
        'material_id': str(gasket), 'qty': '100', 'date': '2026-07-08', 'note': '',
    })
    resp = auth.post(f'/projects/{pid}/stock/issue', data={
        'material_id': str(gasket), 'qty': '150', 'date': '2026-07-08', 'note': '',
    }, follow_redirects=True)
    assert b'blocked' in resp.data
    assert len(query(app, 'SELECT * FROM stock_transactions')) == 1

    auth.post(f'/projects/{pid}/stock/issue', data={
        'material_id': str(gasket), 'qty': '60', 'date': '2026-07-08', 'note': '',
    })
    txn = query_one(app, 'SELECT * FROM stock_transactions ORDER BY id DESC LIMIT 1')
    assert txn['type'] == 'Issue' and txn['qty'] == -60
    cov = _coverage_row(app, auth, pid, 'Gasket')
    assert (cov['received'], cov['issued'], cov['on_hand']) == ('100', '60', '40')


def test_adjustment_requires_reason(app, auth, project):
    # FR-S3 / AC8: adjustments cannot be saved without a reason.
    pid = project['id']
    gasket = material_id(app, 'Gasket')
    auth.post(f'/projects/{pid}/stock/receipt', data={
        'material_id': str(gasket), 'qty': '100', 'date': '2026-07-08', 'note': '',
    })
    resp = auth.post(f'/projects/{pid}/stock/adjust', data={
        'material_id': str(gasket), 'qty': '-5', 'date': '2026-07-08', 'reason': '',
    }, follow_redirects=True)
    assert b'reason is mandatory' in resp.data
    assert len(query(app, 'SELECT * FROM stock_transactions')) == 1

    auth.post(f'/projects/{pid}/stock/adjust', data={
        'material_id': str(gasket), 'qty': '-5', 'date': '2026-07-08',
        'reason': 'breakage during handling',
    })
    txn = query_one(app, 'SELECT * FROM stock_transactions ORDER BY id DESC LIMIT 1')
    assert txn['type'] == 'Adjustment' and txn['qty'] == -5


def test_adjustment_cannot_go_below_zero(app, auth, project):
    pid = project['id']
    gasket = material_id(app, 'Gasket')
    resp = auth.post(f'/projects/{pid}/stock/adjust', data={
        'material_id': str(gasket), 'qty': '-5', 'date': '2026-07-08', 'reason': 'oops',
    }, follow_redirects=True)
    assert b'below zero' in resp.data
    assert query(app, 'SELECT * FROM stock_transactions') == []


def test_stock_ledger_csv_export(app, auth, project):
    pid = project['id']
    gasket = material_id(app, 'Gasket')
    auth.post(f'/projects/{pid}/stock/receipt', data={
        'material_id': str(gasket), 'qty': '480', 'date': '2026-07-08', 'note': 'with windows',
    })
    resp = auth.get(f'/projects/{pid}/stock/export.csv')
    assert resp.mimetype == 'text/csv'
    text = resp.get_data(as_text=True)
    assert 'Gasket' in text and '480' in text and 'Receipt' in text


# --- Reports (§10) ------------------------------------------------------------------

def test_coverage_to_order_gap(app, auth, project, vendor):
    # R2: To-order gap = max(0, Required - Ordered).
    pid = project['id']
    mid = material_id(app, 'Silicone')
    line = add_requirement(auth, pid, mid, 100, approve=True, app=app)
    po = create_po(auth, app, pid, vendor['id'], [line['id']])
    pol = query_one(app, 'SELECT * FROM po_lines WHERE po_id = ?', (po['id'],))
    # Trim the draft PO down to 60 and issue: gap should be 40.
    auth.post(f"/purchase-orders/{po['id']}/lines/{pol['id']}/update",
              data={'qty': '60', 'rate': ''})
    cov = _coverage_row(app, auth, pid, 'Silicone')
    assert cov['ordered'] == '0'  # Draft POs are not "Ordered"
    assert cov['to_order_gap'] == '100'
    auth.post(f"/purchase-orders/{po['id']}/issue")
    cov = _coverage_row(app, auth, pid, 'Silicone')
    assert cov['ordered'] == '60'
    assert cov['to_order_gap'] == '40'


def test_tostem_requisition_report(app, auth, project):
    # AC5: gasket quantity and window summary on the requisition.
    pid = project['id']
    add_requirement(auth, pid, material_id(app, 'Gasket'), 480, note='typed manually', app=app)
    auth.post(f'/projects/{pid}/windows/add', data={
        'opening_ref': 'W-101', 'config_type_id': str(_config_id(app)),
        'width_mm': '2400', 'height_mm': '1500', 'quantity': '3',
    })
    text = auth.get(f'/projects/{pid}/reports/tostem').get_data(as_text=True)
    assert 'Gasket' in text and '480' in text
    # Window summary total row: 1 opening line, 3 windows.
    assert re.search(r'Total</td>\s*<td class="num"[^>]*>1</td>\s*<td class="num"[^>]*>3</td>', text)
    # Vendor-procured materials never appear on the requisition.
    add_requirement(auth, pid, material_id(app, 'Silicone'), 40, app=app)
    text = auth.get(f'/projects/{pid}/reports/tostem').get_data(as_text=True)
    assert 'Silicone' not in text


def test_requirement_report_lists_all_lines(app, auth, project):
    pid = project['id']
    add_requirement(auth, pid, material_id(app, 'Glass'), 120, note='20 windows lot 1', app=app)
    # Safety Shoes is now a Tool -> added via the Tools tab, still shown in the report.
    add_tool(auth, pid, material_id(app, 'Safety Shoes'), 6, note='crew of 6', approve=True, app=app)
    text = auth.get(f'/projects/{pid}/reports/requirements').get_data(as_text=True)
    assert 'Glass' in text and 'Safety Shoes' in text
    assert '20 windows lot 1' in text and 'crew of 6' in text
    assert 'Draft' in text and 'Approved' in text


def test_po_register_filters(app, auth, project, vendor):
    line = add_requirement(auth, project['id'], material_id(app, 'Glass'), 20,
                           approve=True, app=app)
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    text = auth.get('/purchase-orders').get_data(as_text=True)
    assert po['po_number'] in text
    text = auth.get('/purchase-orders?status=Issued').get_data(as_text=True)
    assert po['po_number'] not in text
    text = auth.get(f"/purchase-orders?vendor_id={vendor['id']}&status=Draft").get_data(as_text=True)
    assert po['po_number'] in text


# --- General (FR-G2, FR-G3, FR-R4) ----------------------------------------------------

def test_project_archive_soft_delete(app, auth, project):
    # FR-G3: archive rather than hard delete.
    pid = project['id']
    auth.post(f'/projects/{pid}/archive')
    assert query_one(app, 'SELECT archived FROM projects WHERE id = ?', (pid,))['archived'] == 1
    text = auth.get('/projects/').get_data(as_text=True)
    assert project['name'] not in text
    text = auth.get('/projects/?show_archived=1').get_data(as_text=True)
    assert project['name'] in text


def test_timestamps_present(app, auth, project):
    # FR-G2: records carry created/updated timestamps.
    row = query_one(app, 'SELECT created_at, updated_at FROM projects WHERE id = ?',
                    (project['id'],))
    assert row['created_at'] and row['updated_at']


def test_fr_r4_no_prefilled_quantities(app, auth, project):
    # AC7: no screen pre-computes a material quantity from window dimensions.
    pid = project['id']
    auth.post(f'/projects/{pid}/windows/add', data={
        'opening_ref': 'W-101', 'config_type_id': str(_config_id(app)),
        'width_mm': '2400', 'height_mm': '1500', 'quantity': '10',
    })
    text = auth.get(f'/projects/{pid}/requirements/').get_data(as_text=True)
    # The required-qty input renders empty — never derived from the 10 windows.
    assert 'name="required_qty" required inputmode="decimal" autocomplete="off"' in text
    assert 'value=' not in text.split('required_qty')[1].split('>')[0]


def test_backup_command(app, tmp_path):
    from mpapp.backup import perform_backup
    dest = perform_backup(app.config['DATABASE'], app.config['BACKUP_DIR'])
    import sqlite3
    conn = sqlite3.connect(dest)
    count = conn.execute('SELECT COUNT(*) FROM materials').fetchone()[0]
    conn.close()
    assert count == 9


# ============================================================================
# Drawing-sheet extension: Materials/Tools split, glass parser, formula calc
# ============================================================================

import os

import pytest

pdf_available = os.path.exists(DRAWING_PDF)


# --- Materials / Tools split -------------------------------------------------

def test_seed_split_material_vs_tool(app):
    tools = query(app, "SELECT name FROM materials WHERE item_type = 'Tool' ORDER BY name")
    mats = query(app, "SELECT name FROM materials WHERE item_type = 'Material' ORDER BY name")
    assert {t['name'] for t in tools} == {'Safety Shoes', 'Helmets', 'Jackets'}
    # Masking Tape stays a Material, not a Tool (explicit business rule).
    mat_names = {m['name'] for m in mats}
    assert 'Masking Tape' in mat_names
    assert 'Safety Shoes' not in mat_names


def test_materials_tab_excludes_tools(app, auth, project):
    text = auth.get(f"/projects/{project['id']}/requirements/").get_data(as_text=True)
    assert 'Masking Tape' in text
    assert 'Safety Shoes' not in text  # tools not offered on the materials add form


def test_tools_tab_excludes_materials(app, auth, project):
    text = auth.get(f"/projects/{project['id']}/tools/").get_data(as_text=True)
    assert 'Safety Shoes' in text and 'Helmets' in text and 'Jackets' in text
    assert 'Masking Tape' not in text


def test_material_route_rejects_tool(app, auth, project):
    shoes = material_id(app, 'Safety Shoes')
    resp = auth.post(f"/projects/{project['id']}/requirements/add", data={
        'material_id': str(shoes), 'required_qty': '6', 'basis_note': '',
    }, follow_redirects=True)
    assert b'Pick a material from the catalog' in resp.data
    assert query(app, 'SELECT * FROM requirement_lines') == []


def test_tool_lines_flow_through_coverage_and_po(app, auth, project, vendor):
    shoes = material_id(app, 'Safety Shoes')
    line = add_tool(auth, project['id'], shoes, 6, approve=True, app=app)
    # Tools reuse PO machinery (they are Vendor-procured).
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    pol = query_one(app, 'SELECT * FROM po_lines WHERE po_id = ?', (po['id'],))
    assert pol['qty'] == 6


def test_add_tool_via_catalog(app, auth):
    auth.post('/materials/new', data={
        'name': 'Cordless Drill', 'category': 'Tools', 'uom': 'nos',
        'supply_source': 'Vendor-procured', 'item_type': 'Tool', 'default_vendor_id': '', 'notes': '',
    })
    row = query_one(app, "SELECT * FROM materials WHERE name = 'Cordless Drill'")
    assert row is not None and row['item_type'] == 'Tool'


# --- Glass parser (unit) -----------------------------------------------------

@pytest.mark.skipif(not pdf_available, reason='sample drawing sheet not present')
def test_parser_matches_expected_output():
    from mpapp.glass_parser import parse_pdf_bytes
    with open(DRAWING_PDF, 'rb') as f:
        result = parse_pdf_bytes(f.read())
    assert result['glass_total_qty'] == 916          # exact match to the supplied answer key
    assert len(result['glass']) == 112
    assert result['pages_format_a'] == 5
    assert result['pages_format_b'] == 6
    # Format A anchor: L1W23_60 -> 13 sets x Qty/1set 2 = 26, and x1 = 13.
    a = [g for g in result['glass'] if g['ref'] == 'L1W23_60']
    assert {(g['gw'], g['gh'], g['qty']) for g in a} == {(500, 1203, 26), (1068, 524, 13)}
    # Format B anchor: L1-W7_10 -> Qty column is final (33), no multiplication.
    b = [g for g in result['glass'] if g['ref'] == 'L1-W7_10']
    assert b and b[0]['qty'] == 33 and b[0]['fmt'] == 'B'


def test_parser_rejects_non_drawing_pdf():
    from mpapp.glass_parser import parse_pdf_bytes
    import pytest as _pytest
    with _pytest.raises(ValueError):
        # minimal valid PDF with no tables
        minimal = (b'%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n'
                   b'2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n'
                   b'3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 200 200]>>endobj\n'
                   b'trailer<</Root 1 0 R>>')
        parse_pdf_bytes(minimal)


# --- Formula calculations (unit) --------------------------------------------

def test_formula_reference_anchors():
    from mpapp.services import compute_formula_material
    sil = compute_formula_material('silicone', [{'product_width': 2000, 'product_height': 2000, 'window_qty': 1}])
    assert abs(sil['raw'] - 3.5) < 1e-6           # 2000x2000, 5mm gap -> 3.5 cartridges
    scr = compute_formula_material('screws', [{'product_width': 3000, 'product_height': 3000, 'window_qty': 1}])
    assert abs(scr['raw'] - 80.0) < 1e-6          # 3000x3000, 150mm -> 80 screws
    tap = compute_formula_material('masking_tape', [{'product_width': 2000, 'product_height': 2000, 'window_qty': 1}])
    assert abs(tap['raw'] - 0.8) < 1e-6           # 2*8000/20000


def test_formula_params_scale_live():
    from mpapp.services import compute_formula_material
    prods = [{'product_width': 2000, 'product_height': 2000, 'window_qty': 1}]
    # Doubling the gap doubles the silicone; halving the spacing doubles the screws.
    base_sil = compute_formula_material('silicone', prods, {'gap_mm': 5})['raw']
    dbl_sil = compute_formula_material('silicone', prods, {'gap_mm': 10})['raw']
    assert abs(dbl_sil - 2 * base_sil) < 1e-6
    base_scr = compute_formula_material('screws', prods, {'spacing_mm': 150})['raw']
    tight_scr = compute_formula_material('screws', prods, {'spacing_mm': 75})['raw']
    assert abs(tight_scr - 2 * base_scr) < 1e-6


def test_formula_aggregates_over_windows():
    from mpapp.services import compute_formula_material
    prods = [{'product_width': 2000, 'product_height': 2000, 'window_qty': 3}]
    sil = compute_formula_material('silicone', prods)
    assert abs(sil['raw'] - 3 * 3.5) < 1e-6


# --- Glass sheet integration -------------------------------------------------

@pytest.mark.skipif(not pdf_available, reason='sample drawing sheet not present')
def test_glass_upload_and_formula_end_to_end(app, auth, project):
    pid = project['id']
    with open(DRAWING_PDF, 'rb') as f:
        content = f.read()
    resp = auth.post(f'/projects/{pid}/glass/upload',
                     data={'file': (io.BytesIO(content), 'TG_GROUP_GLASS_SIZES_1.pdf')},
                     content_type='multipart/form-data', follow_redirects=True)
    assert b'Parsed 112 glass line' in resp.data
    sheet = query_one(app, 'SELECT * FROM drawing_sheets WHERE project_id = ?', (pid,))
    assert sheet['glass_total_qty'] == 916
    assert query_one(app, 'SELECT COUNT(*) AS c FROM glass_lines WHERE project_id = ?', (pid,))['c'] == 112
    products = query(app, 'SELECT COUNT(*) AS c FROM drawing_products WHERE project_id = ?', (pid,))
    assert products[0]['c'] > 0

    # Bulk spec applies to all glass lines.
    auth.post(f'/projects/{pid}/glass/bulk-spec', data={
        'thickness': '6mm+1.52pvb+6mm', 'glass_type': 'Clear Laminated', 'glass_color': 'Clear',
    })
    speced = query(app, "SELECT COUNT(*) AS c FROM glass_lines WHERE project_id = ? AND thickness IS NOT NULL", (pid,))
    assert speced[0]['c'] == 112

    # Save formula materials -> requirement lines created with params in the note.
    auth.post(f'/projects/{pid}/requirements/formulas/save', data={
        'methods': ['silicone', 'screws', 'masking_tape'], 'gap_mm': '5', 'spacing_mm': '150',
    })
    autos = {r['calc_method']: r for r in query(
        app, "SELECT rl.*, m.name AS mname FROM requirement_lines rl JOIN materials m ON m.id = rl.material_id"
             " WHERE rl.project_id = ? AND rl.calc_method != 'manual'", (pid,))}
    assert set(autos) == {'silicone', 'screws', 'masking_tape'}
    assert autos['silicone']['required_qty'] > 0
    assert '5mm gap' in (autos['silicone']['basis_note'] or '') or 'gap' in (autos['silicone']['basis_note'] or '')

    # Changing the gap and re-saving updates the Draft silicone line (live recompute).
    before = autos['silicone']['required_qty']
    auth.post(f'/projects/{pid}/requirements/formulas/save', data={
        'methods': ['silicone'], 'gap_mm': '10', 'spacing_mm': '150',
    })
    after = query_one(app, "SELECT required_qty FROM requirement_lines WHERE project_id = ? AND calc_method = 'silicone'", (pid,))['required_qty']
    assert after > before  # bigger gap -> more silicone


@pytest.mark.skipif(not pdf_available, reason='sample drawing sheet not present')
def test_approved_formula_line_not_overwritten(app, auth, project):
    pid = project['id']
    with open(DRAWING_PDF, 'rb') as f:
        content = f.read()
    auth.post(f'/projects/{pid}/glass/upload',
              data={'file': (io.BytesIO(content), 's.pdf')},
              content_type='multipart/form-data')
    auth.post(f'/projects/{pid}/requirements/formulas/save', data={
        'methods': ['silicone'], 'gap_mm': '5', 'spacing_mm': '150',
    })
    line = query_one(app, "SELECT * FROM requirement_lines WHERE project_id = ? AND calc_method = 'silicone'", (pid,))
    auth.post(f"/projects/{pid}/requirements/{line['id']}/approve")
    locked_qty = query_one(app, 'SELECT required_qty FROM requirement_lines WHERE id = ?', (line['id'],))['required_qty']
    # Re-save with a different gap: approved line must be left untouched.
    resp = auth.post(f'/projects/{pid}/requirements/formulas/save', data={
        'methods': ['silicone'], 'gap_mm': '10', 'spacing_mm': '150',
    }, follow_redirects=True)
    assert b'Approved/locked' in resp.data
    assert query_one(app, 'SELECT required_qty FROM requirement_lines WHERE id = ?', (line['id'],))['required_qty'] == locked_qty


def test_gasket_untouched(app, auth, project):
    # Gaskets remain a normal Tostem-supplied material with manual entry, no calc.
    gasket = query_one(app, "SELECT * FROM materials WHERE name = 'Gasket'")
    assert gasket['supply_source'] == 'Tostem-supplied'
    assert gasket['item_type'] == 'Material'
    line = add_requirement(auth, project['id'], gasket['id'], 480, note='typed manually', app=app)
    assert line['calc_method'] == 'manual' and line['required_qty'] == 480


# ============================================================================
# PO workflow extension: Purchase Qty, branded PO, WhatsApp, Glass XLSX export
# ============================================================================

def test_purchase_qty_drives_po(app, auth, project, vendor):
    """Purchase Qty (not Required Qty) goes on the PO; notes carry over."""
    pid = project['id']
    mid = material_id(app, 'Silicone')
    line = add_requirement(auth, pid, mid, 100, approve=True, app=app)
    # Set a deliberate over-order with a note (allowed on the Approved line).
    auth.post(f"/projects/{pid}/requirements/{line['id']}/purchase", data={
        'purchase_qty': '115', 'purchase_notes': 'extra 15 units — Bangalore site buffer',
    })
    row = query_one(app, 'SELECT * FROM requirement_lines WHERE id = ?', (line['id'],))
    assert row['purchase_qty'] == 115
    assert row['required_qty'] == 100  # unchanged — independent values

    po = create_po(auth, app, pid, vendor['id'], [line['id']])
    pol = query_one(app, 'SELECT * FROM po_lines WHERE po_id = ?', (po['id'],))
    assert pol['qty'] == 115
    assert pol['notes'] == 'extra 15 units — Bangalore site buffer'


def test_purchase_qty_blank_falls_back_to_required(app, auth, project, vendor):
    pid = project['id']
    line = add_requirement(auth, pid, material_id(app, 'Glass'), 50, approve=True, app=app)
    po = create_po(auth, app, pid, vendor['id'], [line['id']])
    pol = query_one(app, 'SELECT * FROM po_lines WHERE po_id = ?', (po['id'],))
    assert pol['qty'] == 50  # no purchase qty entered -> required qty is ordered


def test_purchase_qty_validation_and_clearing(app, auth, project):
    pid = project['id']
    line = add_requirement(auth, pid, material_id(app, 'Glass'), 50, app=app)
    resp = auth.post(f"/projects/{pid}/requirements/{line['id']}/purchase", data={
        'purchase_qty': '-3', 'purchase_notes': '',
    }, follow_redirects=True)
    assert b'positive number' in resp.data
    auth.post(f"/projects/{pid}/requirements/{line['id']}/purchase", data={
        'purchase_qty': '60', 'purchase_notes': 'buffer',
    })
    # Blank clears it back to "order the required qty".
    auth.post(f"/projects/{pid}/requirements/{line['id']}/purchase", data={
        'purchase_qty': '', 'purchase_notes': '',
    })
    row = query_one(app, 'SELECT * FROM requirement_lines WHERE id = ?', (line['id'],))
    assert row['purchase_qty'] is None and row['purchase_notes'] is None


def test_tool_purchase_qty(app, auth, project, vendor):
    pid = project['id']
    shoes = material_id(app, 'Safety Shoes')
    line = add_tool(auth, pid, shoes, 6, approve=True, app=app)
    auth.post(f"/projects/{pid}/tools/{line['id']}/purchase", data={
        'purchase_qty': '8', 'purchase_notes': 'two spares',
    })
    po = create_po(auth, app, pid, vendor['id'], [line['id']])
    pol = query_one(app, 'SELECT * FROM po_lines WHERE po_id = ?', (po['id'],))
    assert pol['qty'] == 8 and pol['notes'] == 'two spares'


def test_company_settings_on_printed_po(app, auth, project, vendor):
    auth.post('/settings/company', data={
        'company_name': 'Selective Systems Pvt Ltd',
        'company_tagline': 'Authorised Tostem Partner',
        'company_address': '12 Industrial Estate, Mumbai 400001',
        'company_gst': '27AAAAA0000A1Z5',
        'company_phone': '+91 98200 00000',
        'company_email': 'purchase@selectivesystems.in',
        'po_terms': 'Goods once sold will not be taken back.\nPayment within 30 days.',
        'po_signatory': 'A. Owner',
    })
    line = add_requirement(auth, project['id'], material_id(app, 'Glass'), 20,
                           approve=True, app=app)
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    text = auth.get(f"/purchase-orders/{po['id']}/print").get_data(as_text=True)
    assert 'Selective Systems Pvt Ltd' in text
    assert '27AAAAA0000A1Z5' in text
    assert '12 Industrial Estate' in text
    assert 'Goods once sold will not be taken back' in text
    assert 'A. Owner' in text
    assert 'logo-placeholder.svg' in text


def test_po_terms_placeholder_when_unset(app, auth, project, vendor):
    line = add_requirement(auth, project['id'], material_id(app, 'Glass'), 20,
                           approve=True, app=app)
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    text = auth.get(f"/purchase-orders/{po['id']}/print").get_data(as_text=True)
    assert 'Terms &amp; Conditions' in text
    assert 'to be added' in text  # placeholder shown until filled in Settings


def test_delivery_address_manual_per_po(app, auth, project, vendor):
    line = add_requirement(auth, project['id'], material_id(app, 'Glass'), 20,
                           approve=True, app=app)
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    auth.post(f"/purchase-orders/{po['id']}/update", data={
        'vendor_id': str(vendor['id']), 'order_date': po['order_date'],
        'expected_delivery': '2026-07-20', 'terms_notes': '',
        'delivery_address': 'Gate 3, Skyline Towers site store, Bengaluru',
    })
    row = query_one(app, 'SELECT * FROM purchase_orders WHERE id = ?', (po['id'],))
    assert row['delivery_address'] == 'Gate 3, Skyline Towers site store, Bengaluru'
    text = auth.get(f"/purchase-orders/{po['id']}/print").get_data(as_text=True)
    assert 'Gate 3, Skyline Towers site store' in text


def test_issue_stamps_po_date(app, auth, project, vendor):
    """The PO date is auto-set to the date of issue (spec §2)."""
    line = add_requirement(auth, project['id'], material_id(app, 'Glass'), 20,
                           approve=True, app=app)
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    # Backdate the draft, then issue: order_date must be stamped to today.
    execute(app, 'UPDATE purchase_orders SET order_date = ? WHERE id = ?',
            ('2026-01-01', po['id']))
    auth.post(f"/purchase-orders/{po['id']}/issue")
    row = query_one(app, 'SELECT * FROM purchase_orders WHERE id = ?', (po['id'],))
    assert row['status'] == 'Issued'
    assert row['order_date'] == date.today().isoformat()


def test_whatsapp_link_on_po(app, auth, project, vendor):
    line = add_requirement(auth, project['id'], material_id(app, 'Silicone'), 40,
                           approve=True, app=app)
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    auth.post(f"/purchase-orders/{po['id']}/update", data={
        'vendor_id': str(vendor['id']), 'order_date': po['order_date'],
        'expected_delivery': '', 'terms_notes': '',
        'delivery_address': 'Site store, Tower B',
    })
    text = auth.get(f"/purchase-orders/{po['id']}").get_data(as_text=True)
    # Vendor phone is '99999' (conftest) -> too short for wa.me; disabled state...
    assert 'no vendor phone' not in text or 'wa.me' in text

    # Give the vendor a real phone; the wa.me link must appear, URL-encoded.
    execute(app, 'UPDATE vendors SET phone = ? WHERE id = ?',
            ('+91 98200 12345', vendor['id']))
    text = auth.get(f"/purchase-orders/{po['id']}").get_data(as_text=True)
    assert 'https://wa.me/919820012345?text=' in text
    assert 'Purchase%20Order%20' + po['po_number'].replace('SS-PO-', 'SS-PO-') in text or \
           'Purchase%20Order' in text
    assert 'Site%20store%2C%20Tower%20B' in text  # delivery address in the message
    assert 'Silicone' in text


def test_whatsapp_message_content(app):
    from mpapp.services import build_po_whatsapp_message, whatsapp_phone_digits
    settings = {'company_name': 'Selective Systems'}
    po = {'po_number': 'SS-PO-2026-007', 'order_date': '2026-07-10',
          'delivery_address': 'Gate 3, Site store', 'expected_delivery': '2026-07-20'}
    vendor = {'name': 'Alpha Traders', 'phone': '+91 98200-12345'}
    lines = [{'material_name': 'Silicone', 'qty': 115, 'uom': 'tubes',
              'notes': 'extra 15 — buffer'}]
    msg = build_po_whatsapp_message(settings, po, vendor, lines)
    assert '*Purchase Order SS-PO-2026-007*' in msg
    assert 'Date: 10-07-2026' in msg
    assert '1. Silicone — 115 tubes (extra 15 — buffer)' in msg
    assert 'Delivery address: Gate 3, Site store' in msg
    assert 'Expected delivery: 20-07-2026' in msg
    assert whatsapp_phone_digits(vendor['phone']) == '919820012345'
    assert whatsapp_phone_digits('99999') is None   # too short
    assert whatsapp_phone_digits(None) is None


def test_no_whatsapp_without_phone(app, auth, project, vendor):
    execute(app, 'UPDATE vendors SET phone = NULL WHERE id = ?', (vendor['id'],))
    line = add_requirement(auth, project['id'], material_id(app, 'Glass'), 20,
                           approve=True, app=app)
    po = create_po(auth, app, project['id'], vendor['id'], [line['id']])
    text = auth.get(f"/purchase-orders/{po['id']}").get_data(as_text=True)
    assert 'wa.me' not in text
    assert 'no vendor phone' in text


# --- Glass Sheet: totals + XLSX export ---------------------------------------

@pytest.mark.skipif(not pdf_available, reason='sample drawing sheet not present')
def test_glass_sheet_totals_and_xlsx_export(app, auth, project):
    pid = project['id']
    with open(DRAWING_PDF, 'rb') as f:
        content = f.read()
    auth.post(f'/projects/{pid}/glass/upload',
              data={'file': (io.BytesIO(content), 'draw.pdf')},
              content_type='multipart/form-data')
    auth.post(f'/projects/{pid}/glass/bulk-spec', data={
        'thickness': '6mm+1.52pvb+6mm', 'glass_type': 'Clear Laminated', 'glass_color': 'Clear',
    })
    # Totals row at the bottom of the glass table.
    text = auth.get(f'/projects/{pid}/glass/').get_data(as_text=True)
    assert 'sqm total glass area' in text
    assert '916' in text

    # Processed XLSX export includes the manual spec columns and totals.
    resp = auth.get(f'/projects/{pid}/glass/export.xlsx')
    assert resp.status_code == 200
    assert 'spreadsheetml' in resp.mimetype
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header[:7] == ['Reference Code', 'Glass Width (mm)', 'Glass Height (mm)', 'Qty',
                          'Thickness', 'Glass Type', 'Colour']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    data_rows = [r for r in rows if r[0] and r[0] != 'TOTAL']
    assert len(data_rows) == 112
    assert all(r[4] == '6mm+1.52pvb+6mm' for r in data_rows)
    total_row = next(r for r in rows if r[0] == 'TOTAL')
    assert total_row[3] == 916

    # The original-PDF download is unchanged and still returns the source file.
    sheet = query_one(app, 'SELECT id FROM drawing_sheets WHERE project_id = ?', (pid,))
    resp = auth.get(f"/projects/{pid}/glass/sheet/{sheet['id']}/download")
    assert resp.data == content


def test_glass_xlsx_export_empty_project(app, auth, project):
    resp = auth.get(f"/projects/{project['id']}/glass/export.xlsx", follow_redirects=True)
    assert b'No glass lines to export' in resp.data
